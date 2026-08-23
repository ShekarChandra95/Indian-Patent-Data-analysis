"""Load stage: persist the transformed tables to sqlite / Excel / CSV.

Every function here is a pure sink: it takes the dict of DataFrames
produced by ``transform.transform()`` and writes it somewhere. None of
them mutate or further derive data, so new output formats can be added
without touching extract.py or transform.py.
"""

from __future__ import annotations

import logging
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

_FONT = "Arial"
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(name=_FONT, bold=True, color="FFFFFF", size=10)
_BODY_FONT = Font(name=_FONT, size=10)
_BORDER = Border(*(Side(style="thin", color="B7B7B7") for _ in range(4)))
_URGENCY_FILLS = {
    "OVERDUE": PatternFill("solid", fgColor="F8CBAD"),
    "DUE_SOON": PatternFill("solid", fgColor="FFE699"),
    "DUE_LATER": PatternFill("solid", fgColor="C6E0B4"),
}


def load_to_sqlite(tables: dict[str, pd.DataFrame], db_path: str | Path) -> None:
    """Write every table to a sqlite database, one table per DataFrame key."""
    db_path = Path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        for name, df in tables.items():
            out = df.copy()
            for col in out.select_dtypes(include=["datetime64[ns]"]).columns:
                out[col] = out[col].dt.strftime("%Y-%m-%d")
            out.to_sql(name, conn, if_exists="replace", index=False)
            logger.info("Wrote table '%s' (%d rows) to %s", name, len(out), db_path)


def load_to_csv(tables: dict[str, pd.DataFrame], output_dir: str | Path) -> None:
    """Write every table to its own CSV file under output_dir/csv/."""
    output_dir = Path(output_dir) / "csv"
    output_dir.mkdir(parents=True, exist_ok=True)
    for name, df in tables.items():
        path = output_dir / f"{name}.csv"
        df.to_csv(path, index=False)
        logger.info("Wrote %s (%d rows)", path, len(df))


def _style_sheet(ws, df: pd.DataFrame, table_name: str, date_cols: tuple[str, ...] = ()) -> None:
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.freeze_panes = "A2"

    date_idx = {df.columns.get_loc(c) + 1 for c in date_cols if c in df.columns}
    for row in df.itertuples(index=False):
        ws.append(list(row))

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _BODY_FONT
            cell.border = _BORDER
            if c in date_idx and cell.value is not None:
                cell.number_format = "dd-mmm-yyyy"

    for c, col_name in enumerate(df.columns, start=1):
        width = min(max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) else 0) + 2, 45)
        ws.column_dimensions[get_column_letter(c)].width = width

    if ws.max_row > 1:
        tab = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(df.columns))}{ws.max_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)


def _urgency_fill_key(value: str) -> str | None:
    if value == "OVERDUE":
        return "OVERDUE"
    if isinstance(value, str) and value.startswith("DUE <=") and "90" not in value.split("<=")[1]:
        return "DUE_SOON"
    if isinstance(value, str) and value.startswith("DUE <="):
        return "DUE_LATER"
    return None


def load_to_excel(tables: dict[str, pd.DataFrame], excel_path: str | Path) -> None:
    """Write a formatted, multi-sheet Excel workbook (values, not formulas).

    Values are baked in deliberately: this file is a generated artifact of
    the pipeline, re-created on every run. The pipeline code -- not
    spreadsheet formulas -- is the source of truth for how the numbers are
    computed, which keeps the logic testable and diffable in git.
    """
    from openpyxl import Workbook

    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    sheet_specs = [
        ("Status Summary", tables["status_summary"], "StatusSummaryTable", ()),
        ("Client Summary", tables["client_summary"], "ClientSummaryTable", ("next_rq_due",)),
        ("Pending RQ Deadlines", tables["pending_rq"], "PendingRQTable", ("filing_date", "rq_due")),
        ("Cases", tables["cases"], "CasesTable", ("case_received_on", "filing_date", "rq_due", "rq_filed")),
    ]

    for title, df, table_name, date_cols in sheet_specs:
        ws = wb.create_sheet(title)
        _style_sheet(ws, df, table_name, date_cols)

    # Highlight urgency on the Pending RQ Deadlines sheet
    ws = wb["Pending RQ Deadlines"]
    urgency_col = tables["pending_rq"].columns.get_loc("urgency") + 1
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=urgency_col).value
        key = _urgency_fill_key(val)
        if key:
            ws.cell(row=r, column=urgency_col).fill = _URGENCY_FILLS[key]

    # Highlight clients with overdue / due-soon RQs on the Client Summary sheet
    ws = wb["Client Summary"]
    overdue_col = tables["client_summary"].columns.get_loc("overdue_rq") + 1
    due_col = tables["client_summary"].columns.get_loc("due_within_window") + 1
    for r in range(2, ws.max_row + 1):
        if (ws.cell(row=r, column=overdue_col).value or 0) > 0:
            ws.cell(row=r, column=overdue_col).fill = _URGENCY_FILLS["OVERDUE"]
        if (ws.cell(row=r, column=due_col).value or 0) > 0:
            ws.cell(row=r, column=due_col).fill = _URGENCY_FILLS["DUE_SOON"]

    wb.move_sheet("Status Summary", offset=-len(wb.sheetnames))
    wb.save(excel_path)
    logger.info("Wrote Excel report to %s", excel_path)


def load_all(tables: dict[str, pd.DataFrame], output_dir: str | Path, formats: tuple[str, ...]) -> None:
    """Dispatch to whichever loaders were requested."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if "sqlite" in formats:
        load_to_sqlite(tables, output_dir / "patent_data.db")
    if "csv" in formats:
        load_to_csv(tables, output_dir)
    if "excel" in formats:
        load_to_excel(tables, output_dir / "patent_due_date_report.xlsx")
